from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from bot.config import BASE_DIR


async def generate_excel(
    orders: list[dict],
    products: list[dict],
    session_id: int,
) -> Path:
    """Generate an Excel file with order details.

    Rows = clients, Columns = products, Cells = quantities.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Sesion {session_id}"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Header row
    ws.cell(row=1, column=1, value="Cliente").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).border = thin_border

    ws.cell(row=1, column=2, value="Empresa").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).border = thin_border

    for col_idx, product in enumerate(products, start=3):
        cell = ws.cell(row=1, column=col_idx, value=product["name"])
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Total column
    total_col = len(products) + 3
    ws.cell(row=1, column=total_col, value="TOTAL").font = header_font
    ws.cell(row=1, column=total_col).fill = header_fill
    ws.cell(row=1, column=total_col).border = thin_border

    # Build product_id -> column index mapping
    product_col = {p["id"]: i + 3 for i, p in enumerate(products)}

    # Data rows
    for row_idx, order in enumerate(orders, start=2):
        ws.cell(row=row_idx, column=1, value=order["client_name"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=order.get("company") or "-").border = thin_border

        row_total = 0
        for item in order["items"]:
            col = product_col.get(item["product_id"])
            if col:
                cell = ws.cell(row=row_idx, column=col, value=item["quantity"])
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                row_total += item["quantity"]

        ws.cell(row=row_idx, column=total_col, value=row_total).border = thin_border

    # Summary row at the bottom
    summary_row = len(orders) + 2
    ws.cell(row=summary_row, column=1, value="TOTAL").font = header_font
    ws.cell(row=summary_row, column=1).fill = header_fill
    ws.cell(row=summary_row, column=1).border = thin_border

    grand_total = 0
    for col_idx, product in enumerate(products, start=3):
        col_total = 0
        for order in orders:
            for item in order["items"]:
                if item["product_id"] == product["id"]:
                    col_total += item["quantity"]
        cell = ws.cell(row=summary_row, column=col_idx, value=col_total)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        grand_total += col_total

    ws.cell(row=summary_row, column=total_col, value=grand_total).font = header_font
    ws.cell(row=summary_row, column=total_col).fill = header_fill
    ws.cell(row=summary_row, column=total_col).border = thin_border

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 10)

    # Save
    export_dir = BASE_DIR / "data" / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    file_path = export_dir / f"pedidos_sesion_{session_id}.xlsx"
    wb.save(str(file_path))
    return file_path
